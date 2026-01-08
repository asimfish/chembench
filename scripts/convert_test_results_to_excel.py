#!/usr/bin/env python3
"""
将测试结果日志文件转换为 Excel 格式
"""

import pandas as pd
import re
from pathlib import Path

def parse_test_results(log_file_path: str):
    """解析测试结果日志文件"""
    
    with open(log_file_path, 'r', encoding='utf-8') as f:
        lines = f.readlines()
    
    # 查找表格开始的位置
    table_start = None
    for i, line in enumerate(lines):
        if '任务名称' in line and '无随机化' in line:
            table_start = i
            break
    
    if table_start is None:
        print("❌ 未找到表格数据")
        return None
    
    # 解析表头
    header_line = lines[table_start].strip()
    # 按竖线分割，并清理空白
    headers = [h.strip() for h in header_line.split('|')]
    headers = [h for h in headers if h]  # 移除空字符串
    
    print(f"表头: {headers}")
    
    # 跳过分隔线
    data_start = table_start + 2
    
    # 解析数据行
    data_rows = []
    for line in lines[data_start:]:
        line = line.strip()
        if not line or line.startswith('=') or line.startswith('-'):
            continue
        
        # 如果是空行或者没有 | 符号，说明表格结束
        if '|' not in line:
            break
        
        # 按竖线分割
        cells = [c.strip() for c in line.split('|')]
        cells = [c for c in cells if c]  # 移除空字符串
        
        if len(cells) != len(headers):
            print(f"⚠️  跳过不完整的行: {line}")
            continue
        
        # 清理百分比符号和空格
        cleaned_cells = []
        for i, cell in enumerate(cells):
            if i == 0:  # 任务名称
                cleaned_cells.append(cell)
            else:  # 百分比数据
                # 移除 % 符号和多余空格
                value = cell.replace('%', '').strip()
                try:
                    # 尝试转换为浮点数
                    cleaned_cells.append(float(value))
                except ValueError:
                    cleaned_cells.append(cell)
        
        data_rows.append(cleaned_cells)
    
    print(f"解析了 {len(data_rows)} 行数据")
    
    # 创建 DataFrame
    df = pd.DataFrame(data_rows, columns=headers)
    
    return df

def create_excel(df: pd.DataFrame, output_path: str):
    """创建 Excel 文件并格式化"""
    
    # 创建 Excel writer
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='测试结果', index=False)
        
        # 获取工作表
        workbook = writer.book
        worksheet = writer.sheets['测试结果']
        
        # 设置列宽
        worksheet.column_dimensions['A'].width = 30  # 任务名称列
        for col in ['B', 'C', 'D', 'E', 'F']:
            worksheet.column_dimensions[col].width = 15
        
        # 格式化表头
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        for cell in worksheet[1]:
            cell.font = Font(bold=True, size=11, color='FFFFFF')
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 格式化数据行
        data_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            # 第一列左对齐
            row[0].alignment = Alignment(horizontal='left', vertical='center')
            # 其他列居中
            for cell in row[1:]:
                cell.alignment = data_alignment
                # 添加百分比格式
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '0.0"%"'
            
            # 添加边框
            for cell in row:
                cell.border = border
        
        # 添加条件格式：高成功率用绿色，低成功率用红色
        from openpyxl.formatting.rule import ColorScaleRule
        
        # 对数据列应用颜色刻度
        for col in ['B', 'C', 'D', 'E', 'F']:
            col_range = f'{col}2:{col}{worksheet.max_row}'
            rule = ColorScaleRule(
                start_type='num', start_value=0, start_color='F8696B',   # 红色
                mid_type='num', mid_value=50, mid_color='FFEB84',        # 黄色
                end_type='num', end_value=100, end_color='63BE7B'        # 绿色
            )
            worksheet.conditional_formatting.add(col_range, rule)
    
    print(f"✅ Excel 文件已创建: {output_path}")

def main():
    import argparse
    
    parser = argparse.ArgumentParser(description='将测试结果日志转换为 Excel')
    parser.add_argument('log_file', nargs='?',
                       default='/home/psibot/chembench/test/test_logs/test_results_20260102_183804.log',
                       help='日志文件路径')
    parser.add_argument('-o', '--output', 
                       help='输出 Excel 文件路径（默认：与日志文件同名）')
    
    args = parser.parse_args()
    
    log_path = Path(args.log_file)
    
    if not log_path.exists():
        print(f"❌ 文件不存在: {log_path}")
        return
    
    # 确定输出路径
    if args.output:
        output_path = args.output
    else:
        output_path = log_path.parent / (log_path.stem + '.xlsx')
    
    print(f"\n{'='*60}")
    print(f"转换测试结果为 Excel")
    print(f"{'='*60}")
    print(f"输入文件: {log_path}")
    print(f"输出文件: {output_path}")
    print(f"{'='*60}\n")
    
    # 解析日志文件
    df = parse_test_results(str(log_path))
    
    if df is None:
        return
    
    # 显示数据预览
    print("\n数据预览:")
    print(df.head(10).to_string())
    print(f"\n总计: {len(df)} 个任务\n")
    
    # 创建 Excel 文件
    create_excel(df, str(output_path))
    
    # 计算统计信息
    print("\n统计信息:")
    for col in df.columns[1:]:
        avg = df[col].mean()
        print(f"  {col}: 平均成功率 = {avg:.1f}%")
    
    print(f"\n✅ 转换完成!")

if __name__ == '__main__':
    main()




