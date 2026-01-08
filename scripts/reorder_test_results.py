#!/usr/bin/env python3
"""
按指定顺序重新排列Excel中的测试结果
"""

import pandas as pd
from pathlib import Path

def reorder_excel(input_path: str, output_path: str = None):
    """按指定顺序重新排列Excel数据"""
    
    # 指定的顺序
    desired_order = [
        "100ml玻璃烧杯",
        "250ml棕色容量瓶",
        "250ml玻璃烧杯",
        "500ml玻璃烧杯",
        "坩埚",
        "棕色试剂瓶大",
        "100ml塑料量筒",
        "100ml玻璃量筒",
        "250ml透明容量瓶",
        "500ml塑料量筒",
        "500ml玻璃量筒",
        "50ml玻璃烧杯",
        "1000ml透明容量瓶",
        "500ml透明容量瓶",
        "具塞锥形瓶",
        "漏斗",
        "透明试剂瓶大",
        "酒精灯",
        "棕色试剂瓶小",
        "透明试剂瓶小",
        "锥形瓶"
    ]
    
    # 读取Excel
    df = pd.read_excel(input_path)
    
    print(f"原始数据: {len(df)} 行")
    print(f"任务名称列表:")
    for i, name in enumerate(df['任务名称'], 1):
        print(f"  {i}. {name}")
    
    # 筛选出ND相关的数据（即没有特殊后缀的任务）
    # 排除带有 "_纯RGB", "_RGB_masked", "_RGB6" 等后缀的任务
    exclude_keywords = ["_纯RGB", "_RGB_masked", "_RGB6", "_RGB"]
    
    df_nd = df[~df['任务名称'].str.contains('|'.join(exclude_keywords), na=False)].copy()
    
    print(f"\n筛选后的ND数据: {len(df_nd)} 行")
    print(f"ND任务列表:")
    for i, name in enumerate(df_nd['任务名称'], 1):
        print(f"  {i}. {name}")
    
    # 创建一个排序键的字典
    order_dict = {name: i for i, name in enumerate(desired_order)}
    
    # 添加排序键列
    df_nd['排序键'] = df_nd['任务名称'].map(order_dict)
    
    # 检查是否有未匹配的任务
    unmatched = df_nd[df_nd['排序键'].isna()]
    if not unmatched.empty:
        print(f"\n⚠️  警告：以下任务在排序列表中未找到：")
        for name in unmatched['任务名称']:
            print(f"  - {name}")
    
    # 按排序键排序
    df_nd_sorted = df_nd.sort_values('排序键').drop('排序键', axis=1).reset_index(drop=True)
    
    print(f"\n重新排序后:")
    for i, name in enumerate(df_nd_sorted['任务名称'], 1):
        print(f"  {i}. {name}")
    
    # 确定输出路径
    if output_path is None:
        input_path_obj = Path(input_path)
        output_path = input_path_obj.parent / (input_path_obj.stem + '_nd_sorted.xlsx')
    
    # 写入Excel并格式化
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df_nd_sorted.to_excel(writer, sheet_name='ND测试结果', index=False)
        
        # 获取工作表
        workbook = writer.book
        worksheet = writer.sheets['ND测试结果']
        
        # 设置列宽
        worksheet.column_dimensions['A'].width = 30  # 任务名称列
        for col in ['B', 'C', 'D', 'E', 'F']:
            worksheet.column_dimensions[col].width = 15
        
        # 格式化
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.formatting.rule import ColorScaleRule
        
        # 表头格式
        header_font = Font(bold=True, size=11, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 数据格式
        data_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            row[0].alignment = Alignment(horizontal='left', vertical='center')
            for cell in row[1:]:
                cell.alignment = data_alignment
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '0.0"%"'
            for cell in row:
                cell.border = border
        
        # 条件格式（颜色渐变）
        for col in ['B', 'C', 'D', 'E', 'F']:
            col_range = f'{col}2:{col}{worksheet.max_row}'
            rule = ColorScaleRule(
                start_type='num', start_value=0, start_color='F8696B',
                mid_type='num', mid_value=50, mid_color='FFEB84',
                end_type='num', end_value=100, end_color='63BE7B'
            )
            worksheet.conditional_formatting.add(col_range, rule)
    
    print(f"\n✅ 重新排序的Excel文件已创建: {output_path}")
    
    # 计算统计信息
    print("\n统计信息:")
    for col in df_nd_sorted.columns[1:]:
        avg = df_nd_sorted[col].mean()
        print(f"  {col}: 平均成功率 = {avg:.1f}%")
    
    return df_nd_sorted

def main():
    import argparse
    
    parser = argparse.ArgumentParser(description='按指定顺序重新排列Excel中的ND测试结果')
    parser.add_argument('input_file', nargs='?',
                       default='/home/psibot/chembench/test/test_logs/test_results_20260102_183804.xlsx',
                       help='输入Excel文件路径')
    parser.add_argument('-o', '--output',
                       help='输出Excel文件路径（默认：*_nd_sorted.xlsx）')
    
    args = parser.parse_args()
    
    input_path = Path(args.input_file)
    
    if not input_path.exists():
        print(f"❌ 文件不存在: {input_path}")
        return
    
    print(f"\n{'='*60}")
    print(f"重新排序测试结果（ND数据）")
    print(f"{'='*60}")
    print(f"输入文件: {input_path}\n")
    
    reorder_excel(str(input_path), args.output)
    
    print(f"\n✅ 完成!")

if __name__ == '__main__':
    main()




